#!/usr/bin/env python3
# -*- coding:utf-8 -*-

import requests
from lxml import etree
import codecs
import time
from contextlib import contextmanager
import shutil
from openpyxl import load_workbook
from pydub import AudioSegment
from pydub.utils import make_chunks
from functools import reduce
import enchant
import re
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
import os
import math
from PIL import Image, ImageFont, ImageDraw, ImageEnhance, ImageChops


class WaterMarker:
    mark_text = '星航翻译工作室'
    input_path = './todo/input_image'
    out_path = './todo/output_image'
    color = '#8B8B1B'
    space = 150
    angle = 30
    size = 200
    opacity = 0.03

    def _add_mark(self, imagePath, mark):
        """
        添加水印，然后保存图片
        """
        im = Image.open(imagePath)

        image = mark(im)
        if image:
            name = os.path.basename(imagePath)
            if not os.path.exists(self.out_path):
                os.mkdir(self.out_path)

            new_name = os.path.join(self.out_path, name)
            if os.path.splitext(new_name)[1] != '.png':
                image = image.convert('RGB')
            image.save(new_name)

    def _set_opacity(self, im, opacity):
        """
        设置水印透明度
        """
        assert 0 <= opacity <= 1

        alpha = im.split()[3]
        alpha = ImageEnhance.Brightness(alpha).enhance(opacity)
        im.putalpha(alpha)
        return im

    def _crop_image(self, im):
        """
        裁剪图片边缘空白
        """
        bg = Image.new(mode='RGBA', size=im.size)
        diff = ImageChops.difference(im, bg)
        del bg
        bbox = diff.getbbox()
        if bbox:
            return im.crop(bbox)
        return im

    def _gen_mark(self):
        """
        生成mark图片，返回添加水印的函数
        """
        # 字体宽度
        width = len(self.mark_text) * self.size

        # 创建水印图片(宽度、高度)
        mark = Image.new(mode='RGBA', size=(width, self.size))

        # 生成文字
        draw_table = ImageDraw.Draw(im=mark)
        draw_table.text(xy=(0, 0), text=self.mark_text, fill=self.color,
                        font=ImageFont.truetype(font_path, size=self.size))
        del draw_table

        # 裁剪空白
        mark = self._crop_image(mark)

        # 透明度
        self._set_opacity(mark, self.opacity)

        # 在im图片上添加水印 im为打开的原图
        def mark_im(im):
            # 计算斜边长度
            c = int(math.sqrt(im.size[0] * im.size[0] + im.size[1] * im.size[1]))
            # 以斜边长度为宽高创建大图（旋转后大图才足以覆盖原图）
            mark2 = Image.new(mode='RGBA', size=(c, c))
            # 在大图上生成水印文字，此处mark为上面生成的水印图片
            y, idx = 0, 0
            while y < c:
                # 制造x坐标错位
                x = -int((mark.size[0] + self.space) * 0.5 * idx)
                idx = (idx + 1) % 2
                while x < c:
                    # 在该位置粘贴mark水印图片
                    mark2.paste(mark, (x, y))
                    x = x + mark.size[0] + self.space
                y = y + mark.size[1] + self.space
            # 将大图旋转一定角度
            mark2 = mark2.rotate(self.angle)
            # 在原图上添加大图水印
            if im.mode != 'RGBA':
                im = im.convert('RGBA')
            # 大图，坐标
            im.paste(mark2, (int((im.size[0] - c) / 2), int((im.size[1] - c) / 2)), mask=mark2.split()[3])
            del mark2
            return im

        return mark_im

    def start(self):
        mark = self._gen_mark()
        names = os.listdir(self.input_path)
        for name in names:
            image_file = os.path.join(self.input_path, name)
            self._add_mark(image_file, mark)


class Util:
    """
    工具类
    """

    @staticmethod
    def is_empty(string):
        """
        空内容检查
        :param string: 字符串
        :return: True | False
        """
        empty = [None, '', '无']
        if string in empty:
            return True
        else:
            # 出现 多个空格 / True / False 时
            return str(string).strip() in empty

    @staticmethod
    def convert_to_string(string, default_string=''):
        """
        转换成字符串
        :param string: 字符串
        :param default_string: 空白字符串时的默认值
        :return: 字符串
        """
        # 去除前换换行/空格符号，并将空内容转换为''字符
        return str(string).strip() if not Util.is_empty(string) else default_string

    @staticmethod
    @contextmanager
    def workbook_manager(file_path):
        """
        用上下文管理器加载文件 with workbook_manager('xx') as workbook
        :param file_path: 文件路径 '.\\GRE.xls'
        """
        # 加载文件
        workbook = load_workbook(file_path, data_only=True)
        try:
            yield workbook
        except RuntimeError as e:
            Util.log('加载文件出错', error=e, mode='error')
        finally:
            # 关闭文件
            workbook.close()

    @staticmethod
    def clock(func):
        """
        计时装饰器
        :param func: 被装饰函数
        """

        def clocked(*args, **kwargs):
            start = time.time()
            result = func(*args, **kwargs)
            end = time.time()
            print('耗时：{}s'.format(round(end - start, 2)))
            return result

        return clocked

    @staticmethod
    def log(content, error=None, mode='info', is_print=True, new_line=True):
        """
        日志记录
        :param content: 内容
        :param error: 错误对象
        :param mode: 记录模式 info/error/warn
        :param is_print: 是否打印到屏幕
        :param new_line: 写入文件时是否追加换行符
        """
        if mode == 'error':
            content = '错误 -> {} {}'.format(content, error)
        with codecs.open(log_path, 'a+', 'utf-8') as f:
            f.write(content)
            if new_line:
                f.write('\r\n')
            if is_print:
                print(content)


def get_word(file_path):
    """
    获取单词
    :param file_path: 文件路径 './GRE.xls'
    """
    with Util.workbook_manager(file_path) as workbook:
        # 加载工作表
        worksheet = workbook.active
        for row in worksheet.iter_rows(min_row=start_row_index, max_col=max_col_index):
            yield Util.convert_to_string(row[default_word_col].value)


@Util.clock
def get_info(file_path):
    """
    获取音标/翻译/例句
    :param file_path: 文件路径 './GRE.xls'
    """
    with Util.workbook_manager(file_path) as workbook:
        # 加载工作表
        worksheet = workbook.active
        # 当前行标
        current_row_index = start_row_index
        for row in worksheet.iter_rows(min_row=start_row_index, max_col=max_col_index):
            # 原表格信息
            index = row[default_index_col].value
            word = Util.convert_to_string(row[default_word_col].value)
            phonetic = Util.convert_to_string(row[default_phonetic_col].value)
            translation = Util.convert_to_string(row[default_translation_col].value)
            example = Util.convert_to_string(row[default_example_col].value)

            # 获取单词信息
            try:
                tip = []
                if Util.is_empty(word):
                    tip.append('单词不存在')
                else:
                    if Util.is_empty(phonetic) or Util.is_empty(translation) or Util.is_empty(example):
                        url = 'http://www.youdao.com/w/eng/{}'.format(word)
                        # 超过20秒，跳过该单词
                        data = requests.get(url, timeout=20).text
                        html = etree.HTML(data)

                        # 音标
                        if Util.is_empty(phonetic):
                            uk_phonetic = html.xpath('//*[@id="phrsListTab"]/h2/div/span[1]/span/text()')
                            us_phonetic = html.xpath('//*[@id="phrsListTab"]/h2/div/span[2]/span/text()')
                            if len(uk_phonetic) > 0 or len(us_phonetic) > 0:
                                phonetic = ''
                                if len(uk_phonetic) > 0:
                                    phonetic += '英{} '.format(uk_phonetic[0].strip())
                                if len(us_phonetic) > 0:
                                    phonetic += '美{} '.format(us_phonetic[0].strip())
                                phonetic = phonetic.replace('[', '/').replace(']', '/').strip()
                            else:
                                tip.append('无音标')
                                phonetic = '无'

                        # 翻译
                        if Util.is_empty(translation):
                            translation_li_list = html.xpath(
                                '//*[@id="phrsListTab"]/*[@class="trans-container"]/ul/li/text()')
                            translation = ' '.join([i.replace(' ', '') for i in translation_li_list]).strip()
                            if Util.is_empty(translation):
                                tip.append('无翻译')
                                translation = '无'

                        # 例句
                        if Util.is_empty(example):
                            example = html.xpath('//div[@id="authority"]/ul[@class="ol"]/li[1]/p[1]')
                            if len(example) > 0:
                                example = example[0].xpath('string(.)').strip()
                            else:
                                tip.append('无例句')
                                example = '无'

                        print('{} {} {} {}'.format(word, phonetic, translation, example))
                        worksheet.cell(row=current_row_index, column=default_phonetic_col + 1).value = phonetic
                        worksheet.cell(row=current_row_index, column=default_translation_col + 1).value = translation
                        worksheet.cell(row=current_row_index, column=default_example_col + 1).value = example
                if len(tip) > 0:
                    tips = ' '.join(tip)
                    Util.log('{}{} -> {} -> {}'.format('*' if len(tip) >= 2 else '', index, word, tips))
            except Exception as e:
                Util.log(word, error=e, mode='error')
            # 下一行
            current_row_index += 1
        workbook.save(file_path)


@Util.clock
def check_word(file_path):
    """
    检查单词(拼写/音标/例句)
    :param file_path: 文件路径 './GRE.xls'
    """
    # 拼写检查库，美国、英国字典
    us_enchant = enchant.Dict('en_US')
    gb_enchant = enchant.Dict('en_GB')
    with Util.workbook_manager(file_path) as workbook:
        # 加载工作表
        worksheet = workbook.active
        # 当前行标
        current_row_index = start_row_index
        for row in worksheet.iter_rows(min_row=start_row_index, max_col=max_col_index):
            index = row[default_index_col].value
            word = Util.convert_to_string(row[default_word_col].value)
            phonetic = Util.convert_to_string(row[default_phonetic_col].value)
            translation = Util.convert_to_string(row[default_translation_col].value)
            example = Util.convert_to_string(row[default_example_col].value)
            tip = []
            try:
                if Util.is_empty(phonetic):
                    tip.append('无音标')
                if Util.is_empty(translation):
                    tip.append('无翻译')
                if Util.is_empty(example):
                    tip.append('无例句')
                # if not us_enchant.check(word):
                #     tip.append('美式拼写疑似错误')
                # if not gb_enchant.check(word):
                #     tip.append('英式拼写疑似错误')
                if len(tip) > 0:
                    tips = '{} -> {} {}'.format(index, word, ' '.join(tip))
                    Util.log(tips)
            except Exception as e:
                Util.log(word, error=e, mode='error')
            current_row_index += 1


@Util.clock
def create_image(file_path, output_images_path, font_path, bg_path='./default_bg.png'):
    """
    制作单词图片
    :param file_path: 文件路径 './GRE.xls'
    :param output_images_path: 输出图片目录路径 './图片'
    :param font_path: 字体路径 './YaHeiMyriadPro-Semibold.ttf'
    :param bg_path: 背景图片路径 './default_bg.png'
    """
    # 单词字体
    font_word = ImageFont.truetype(font_path, 80)
    # 音标字体
    font_phonetic = ImageFont.truetype(font_path, 50)
    # 翻译字体
    font_translation = ImageFont.truetype(font_path, 28)
    # 图片大小
    image_size = (1920, 1080)

    # 不存在背景图，创建渐变色背景图
    def create_bg():
        _image = Image.new(mode='RGB', size=image_size, color='white')
        _draw = ImageDraw.Draw(im=_image)
        bg_1 = (17, 153, 142)
        bg_2 = (56, 239, 125)
        step_r = (bg_2[0] - bg_1[0]) / image_size[0]
        step_g = (bg_2[1] - bg_1[1]) / image_size[0]
        step_b = (bg_2[2] - bg_1[2]) / image_size[0]
        for y in range(0, image_size[0] + 1):
            bg_r = round(bg_1[0] + step_r * y)
            bg_g = round(bg_1[1] + step_g * y)
            bg_b = round(bg_1[2] + step_b * y)
            for x in range(0, image_size[0]):
                _draw.point((x, y), fill=(bg_r, bg_g, bg_b))
        _image.save(bg_path, 'PNG')
        # _image.show()
        _image.close()

    if not os.path.isfile(bg_path):
        create_bg()

    with Util.workbook_manager(file_path) as workbook:
        # 加载工作表
        worksheet = workbook.active
        # 当前行标
        current_row_index = start_row_index
        for row in worksheet.iter_rows(min_row=start_row_index, max_col=max_col_index):
            index = row[default_index_col].value
            word = Util.convert_to_string(row[default_word_col].value)
            phonetic = Util.convert_to_string(row[default_phonetic_col].value)
            translation = Util.convert_to_string(row[default_translation_col].value)
            # 翻译长度超过图片宽度，终止生成
            translation_width = font_translation.getsize(translation)
            if translation_width[0] > image_size[0]:
                Util.log(word,
                         error='行标：{}，翻译长度：{}，超过图片宽度，建议手动给翻译换行，然后重新运行生成。'.format(current_row_index, translation_width),
                         mode='error')
                exit()
            # 输出图片路径
            output_image_path = '{}/{}.{}.png'.format(output_images_path, index, word)
            shutil.copyfile(bg_path, output_image_path)
            image = Image.open(output_image_path)
            draw = ImageDraw.Draw(im=image)
            draw.text(xy=(860, 240), text=u'{}'.format(word), fill='#FFFFFF', font=font_word)
            draw.text(xy=(865, 340), text=u'{}'.format(phonetic), fill='#FFFFFF', font=font_phonetic)
            draw.text(xy=(20, 500), text=u'{}'.format(translation), fill='#FFFFFF', font=font_translation)
            # image.show()
            image.save(output_image_path, 'PNG')
            image.close()
            print(current_row_index, word)
            current_row_index += 1


@Util.clock
def get_audio(file_path, audio_path, suffix, speech_type=1):
    """
    获取音频
    :param file_path: 文件路径 './GRE.xls'
    :param audio_path: 音频路径 './音频'
    :param suffix: 音频后缀 '.mp3'
    :param speech_type: 读音类型 英音1/美音2
    """
    word_list = [w.strip() for w in get_word(file_path)]
    for word in word_list:
        # 单词音频是否存在
        word_audio_path = '{}/{}{}'.format(audio_path, word, suffix)
        if not os.path.isfile(word_audio_path):
            try:
                url = 'https://dict.youdao.com/dictvoice?audio={}&type={}'.format(word, speech_type)
                # 超过20秒，跳过该单词
                response = requests.get(url, timeout=20)
                # 保存音频
                with open(word_audio_path, 'wb') as f:
                    f.write(response.content)
                print('{} -> {}'.format(word, response.status_code == 200))
            except Exception as e:
                Util.log(word, error=e, mode='error')


@Util.clock
def find_audio_without_word(file_path, audio_path, suffix):
    """
    输出未找到单词的音频
    :param file_path: 文件路径 './GRE.xls'
    :param audio_path: 音频目录路径 './音频'
    :param suffix: 后缀 '.mp3'
    """
    # 音频名称列表
    audio_name_list = []
    for root, dirs, files in os.walk(audio_path):
        for file in files:
            audio_name_list.append(file.replace(suffix, ''))
    # 单词列表
    word_list = [w.strip() for w in get_word(file_path)]
    # 音频列表有 && 单词列表也有
    a = [w for w in audio_name_list if w in word_list]
    # 合并两个列表，并找出音频列表中没有的
    b = [w for w in (audio_name_list + word_list) if w not in a]
    print('[{}]音频列表有 && 单词列表也有：{}'.format(len(a), a))
    print('[{}]音频列表中没有的单词：{}'.format(len(b), b))


@Util.clock
def remark_audio(file_path, audio_path, suffix, output_audio_path):
    """
    标注音频序号，方便剪辑软件直接排序导入
    :param file_path: 文件路径 './GRE.xls'
    :param audio_path: 音频目录路径 './音频'
    :param suffix: 后缀 '.mp3'
    :param output_audio_path: 输出音频路径 './音频2'
    """
    # 音频列表
    audio_list = []
    for root, dirs, files in os.walk(audio_path):
        for file in files:
            audio_list.append(file)
    # 单词列表
    word_list = [w.strip() for w in get_word(file_path)]
    number = 1
    for word in word_list:
        for audio in audio_list:
            # 找到该单词音频
            audio_name = audio.replace(suffix, '')
            if word == audio_name:
                # 音频文件名称打上序号
                remark = '{}.{}'.format(number, audio)
                remark_path = '{}/{}'.format(output_audio_path, remark)
                with open(remark_path, 'wb') as output_file:
                    with open('{}/{}'.format(audio_path, audio), 'rb') as input_file:
                        output_file.write(input_file.read())
                print('{} -> {}'.format(audio_name, remark))
                number += 1


def sound_slice_normalize(sound, target_level, sample_rate=3 * 1000):
    """
    统一响度
    :param sound: 音频
    :param target_level: 目标响度
    :param sample_rate: 切片长度
    """
    if len(sound) < sample_rate:
        return sound

    set_to_target_level = lambda sound, target_level: sound.apply_gain(target_level - sound.dBFS)

    def max_min_volume(min_level, max_level):
        for chunk in make_chunks(sound, sample_rate):
            if chunk.dBFS < min_level:
                yield set_to_target_level(chunk, min_level)
            elif chunk.dBFS > max_level:
                yield set_to_target_level(chunk, max_level)
            else:
                yield chunk

    return reduce(lambda x, y: x + y, max_min_volume(target_level['min'], target_level['max']))


def combine_all_audio(output_audio_path, combine_audio):
    """
    合并音频
    :param output_audio_path: 输出路径
    :param combine_audio: 最后输出音频名称
    pydub 合并两个音频，需要设置 ffmpeg 环境变量
    https://github.com/jiaaro/pydub/blob/master/API.markdown
    """
    audio_list = []
    for root, dirs, files in os.walk(output_audio_path):
        for file in files:
            audio_list.append(file)
    # 按数字排序
    audio_list.sort(key=lambda x: int(re.match(r'(\d+)', x).group()))

    combined_audio_path = '{}/{}'.format(output_audio_path, combine_audio)

    # 1s 空白音频
    one_second_silence_audio = AudioSegment.silent(duration=1000)
    if not os.path.isfile(combined_audio_path):
        one_second_silence_audio.export(combined_audio_path)

    number = 1
    for audio in audio_list:
        try:
            combined_audio = AudioSegment.from_mp3(combined_audio_path)
            next_audio = AudioSegment.from_mp3('{}/{}'.format(output_audio_path, audio))

            # 获取两个音频的时长，单位：秒
            # combined_audio_time = len(combined_audio) / 1000
            # next_audio_time = len(next_audio) / 1000

            # 获取两个音频的响度
            # combined_audio_db = combined_audio.dBFS
            # next_audio_db = next_audio.dBFS
            # print('第{}次合并音频 {} 响度：{}，时长：{}s'.format(number, combine_audio, combined_audio_db, combined_audio_time))
            # print('当前音频 {} 响度：{}，时长：{}s'.format(audio, next_audio_db, next_audio_time))

            print('第{}次合并 -> {}'.format(number, audio))
            # 合并音频，中间插入空白音频
            new_combined_audio = combined_audio + one_second_silence_audio + next_audio
            new_combined_audio.export(combined_audio_path)
        except Exception as e:
            Util.log(audio, error=e, mode='error')
        number += 1

    # 最后统一响度
    combined_audio = AudioSegment.from_mp3(combined_audio_path)
    new_combined_audio = sound_slice_normalize(combined_audio, normalized_db)
    new_combined_audio.export(combined_audio_path)


@Util.clock
def compare(file_path, file_path2):
    """
    对比单词
    :param file_path: 文件路径 './GRE.xls'
    :param file_path2: 文件2路径 './GRE2.xls'
    """
    word_list = [w.strip() for w in get_word(file_path)]
    word_set = set(word_list)
    word_list2 = [w.strip() for w in get_word(file_path2)]
    word_set2 = set(word_list2)

    same_word = word_set & word_set2
    difference_in_set1 = word_set.difference(word_set2)
    difference_in_set2 = word_set2.difference(word_set)

    print('表1：{}个，表2：{}个'.format(len(word_set), len(word_set2)), end='')
    # print('，相同{}个：{}'.format(len(same_word), same_word))
    print('表1特有{}个：{}'.format(len(difference_in_set1), difference_in_set1))
    print('表2特有{}个：{}'.format(len(difference_in_set2), difference_in_set2))


@Util.clock
def re_make(file_path, not_only_order=True):
    """
    注意：进行该步前请先补全单词信息，以免手动修改后又影响表格顺序
    重新编号。删掉前后空格，英音美音分离
    :param file_path: 文件路径 './GRE.xls'
    :param not_only_order: 不只是编号，还有后面的处理
    """
    with Util.workbook_manager(file_path) as workbook:
        worksheet = workbook.active
        current_row_index = start_row_index

        if not_only_order:
            # 设置样式
            black_color = '000000'
            white_color = 'FFFFFF'
            header_font = Font('微软雅黑', size=12, bold=True, color=white_color)
            font = Font('微软雅黑', size=12, color=black_color)
            header_fill = PatternFill('solid', fgColor='FF5722')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            alignment = Alignment(vertical='center', wrap_text=True)
            header_border = Border(left=Side(border_style='thin', color=white_color),
                                   right=Side(border_style='thin', color=white_color),
                                   top=Side(border_style='thin', color=white_color),
                                   bottom=Side(border_style='thin', color=white_color))
            border = Border(left=Side(border_style='thin', color=black_color),
                            right=Side(border_style='thin', color=black_color),
                            top=Side(border_style='thin', color=black_color),
                            bottom=Side(border_style='thin', color=black_color))

            worksheet.cell(row=1, column=default_phonetic_col + 1).value = '英式音标'
            worksheet.cell(row=1, column=max_col_index).value = '美式音标'
            worksheet['J1'].font = header_font
            worksheet['J1'].fill = header_fill
            worksheet['J1'].border = header_border
            worksheet['J1'].alignment = header_alignment

        for row in worksheet.iter_rows(min_row=start_row_index, max_col=max_col_index):
            worksheet.cell(row=current_row_index, column=default_index_col + 1).value = current_row_index - 1

            if not_only_order:
                worksheet.cell(row=current_row_index, column=default_word_col + 1).value = Util.convert_to_string(
                    row[default_word_col].value)
                worksheet.cell(row=current_row_index, column=default_phonetic_col + 1).value = Util.convert_to_string(
                    row[default_phonetic_col].value)
                worksheet.cell(row=current_row_index,
                               column=default_translation_col + 1).value = Util.convert_to_string(
                    row[default_translation_col].value)
                worksheet.cell(row=current_row_index, column=default_example_col + 1).value = Util.convert_to_string(
                    row[default_example_col].value)

                phonetic = Util.convert_to_string(row[default_phonetic_col].value)
                if '英' in phonetic and '美' in phonetic:
                    uk_phonetic = phonetic.split('美')[0].replace('英', '').strip()
                    us_phonetic = phonetic.split('美')[1].replace('美', '').strip()
                    worksheet.cell(row=current_row_index, column=default_phonetic_col + 1).value = uk_phonetic
                    worksheet.cell(row=current_row_index, column=max_col_index).value = us_phonetic
                else:
                    if '英' in phonetic:
                        uk_phonetic = phonetic.replace('英', '').strip()
                        worksheet.cell(row=current_row_index, column=default_phonetic_col + 1).value = uk_phonetic
                    elif '美' in phonetic:
                        us_phonetic = phonetic.replace('美', '').strip()
                        worksheet.cell(row=current_row_index, column=default_phonetic_col + 1).value = us_phonetic
                worksheet['J{}'.format(current_row_index)].font = font
                worksheet['J{}'.format(current_row_index)].alignment = alignment
                worksheet['J{}'.format(current_row_index)].border = border

            current_row_index += 1
        workbook.save(file_path)


def handle_dir_all(input_dir_path):
    """
    对文件夹内的文件全部处理
    :param input_dir_path: 文件夹
    """
    for root, dirs, files in os.walk(input_dir_path):
        for file in files:
            file_path = '{}/{}'.format(input_dir_path, file)
            global log_path
            log_path = './log/{}.log'.format(file)
            re_make(file_path, False)


if __name__ == '__main__':
    # 开始行标，默认从第2行开始读
    start_row_index = 2
    # 最大10列，避免新数据覆盖旧数据
    max_col_index = 10
    # 默认序号列
    default_index_col = 0
    # 默认单词列
    default_word_col = 1
    # 默认音标列
    default_phonetic_col = 2
    # 默认翻译列
    default_translation_col = 3
    # 默认例句列
    default_example_col = 4

    log_path = './Excel.log'

    # handle_dir_all('./before')
    # exit()

    # 目标文件名
    file_name = '0.xlsx'
    file_name2 = '【正序】四级.xlsx'
    # 单词文件，必须为 xlsx 格式
    file_path = './todo/{}'.format(file_name)
    file_path2 = './todo/{}'.format(file_name2)

    if not os.path.isfile(file_path):
        print('单词文件不存在')
        exit(1)

    # 输出图片路径
    output_images_path = './todo/image'
    # 背景图路径
    bg_path = './resources/image/default_bg.png'
    # 字体路径
    font_path = './resources/font/YaHeiMyriadPro-Semibold.ttf'

    if not os.path.isfile(font_path):
        print('字体文件不存在，加载默认字体。注意：默认字体可能不带中文库')
        font_path = ImageFont.load_default()

    # 音频路径
    audio_path = './todo/audio'
    # 音频后缀
    suffix = '.mp3'
    # 输出音频路径
    output_audio_path = './todo/afterAudio'
    # 合并音频名称
    combine_audio = 'ALL.mp3'
    # 最大最小响度
    normalized_db = {
        'max': -11.0,
        'min': -13.0
    }

    while True:
        print('       1.读取单词                      2.获取音标/翻译/例句')
        print('       3.检查单词(拼写/音标/例句)       4.制作单词图片')
        print('       5.获取音频(英音1/美音2)          6.输出未找到单词的音频')
        print('       7.标注音频序号                   8.合并音频音频')
        print('       9.对比单词                       10.重新编号，删掉前后空格，英音美音分离')
        print('       11.给图片打水印')
        print('       0.退出')
        cmd = input('请输入您需要的功能序号：')
        if cmd == '0':
            exit()
        if cmd == '1':
            for i in get_word(file_path):
                print(i)
        elif cmd == '2':
            get_info(file_path)
        elif cmd == '3':
            check_word(file_path)
        elif cmd == '4':
            create_image(file_path, output_images_path, font_path)
        elif cmd == '5' or cmd == '51':
            get_audio(file_path, audio_path, suffix)
        elif cmd == '52':
            get_audio(file_path, audio_path, suffix, 2)
        elif cmd == '6':
            find_audio_without_word(file_path, audio_path, suffix)
        elif cmd == '7':
            remark_audio(file_path, audio_path, suffix, output_audio_path)
        elif cmd == '8':
            combine_all_audio(output_audio_path, combine_audio)
        elif cmd == '9':
            compare(file_path, file_path2)
        elif cmd == '10':
            re_make(file_path)
        elif cmd == '11':
            WaterMarker().start()
