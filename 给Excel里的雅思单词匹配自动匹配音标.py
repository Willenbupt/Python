import requests
from lxml import etree
from openpyxl import load_workbook


def run():
    #首先存入单词表的路径
    file = 'E:\店铺\单词\雅思\1.xlsx'
    
    #打开目标单词表的excel文件，此时就可以把workbook当作一个excel表对象来使用
    workbook = load_workbook(file)
    
    #读取表里面的'常用9000'那张表
    worksheet = workbook['常用9400']
    
    #设置初始单元格
    row_index = 2
    
    #iter_rows()可以获得多个单元格，最低是第二行，最高是第四列 
    for row in worksheet.iter_rows(min_row=2, max_col=4):
        #获取单词的值
        word = row[1].value
        
        #format是格式化函数，可以将 变量word的值传进{}。
        url = 'http://www.youdao.com/w/eng/{}'.format(word)
        try:
            #获取该网页的HTML最通常的方法是通过r=request.get（url）构造一个向服务器请求资源的url对象。
            #这个对象是Request库内部生成的。
            #这时候的r返回的是一个包含服务器资源的Response对象。包含从服务器返回的所有的相关资源。
            #并保存其text数据
            data = requests.get(url).text
            
            #etree.HTML():构造了一个XPath解析对象并对HTML文本进行自动修正。
            html = etree.HTML(data)
            
            #根据经过修正的HTML文本，定位音标所在的标签位置，并将其返回
            yinbiao = html.xpath('//*[@id="phrsListTab"]/h2/div/span[1]/span/text()')[0]
            print(yinbiao)
            
            #引用音标所在的行并把音标传进excel表里对应的单词
            worksheet.cell(row=row_index, column=4).value = yinbiao
        except Exception as e:
            print(e, word)
        #行加一
        row_index += 1
    workbook.save(file)


if __name__ == '__main__':
    run()
