#!/usr/bin/env python3
# -*- coding:utf-8 -*-

import json
import os
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side


def re_order_index(input_dir_path):
    for root, dirs, files in os.walk(input_dir_path):
        for file in files:
            file_path = '{}\\{}'.format(input_dir_path, file)
            # 加载文件
            workbook = load_workbook(file_path, data_only=True)
            try:
                worksheet = workbook.active
                current_row_index = start_row_index
                for _ in worksheet.iter_rows(min_row=start_row_index, max_col=10):
                    worksheet.cell(row=current_row_index, column=default_index_col).value = current_row_index - 1
                    current_row_index += 1
                workbook.save(file_path)
            except RuntimeError as e:
                print('加载文件出错', e)
            finally:
                workbook.close()


def parse_all(input_dir_path, output_dir_path):
    # 读取文件夹下的 json 数据
    for root, dirs, files in os.walk(input_dir_path):
        for file in files:
            file_path = '{}\\{}'.format(input_dir_path, file)

            # 写入 Excel
            xlsx_path = '{}\\{}.xlsx'.format(output_dir_path, file)
            workbook = openpyxl.Workbook()
            # 获取活跃的工作表
            worksheet = workbook.active
            # 当前行标
            current_row_index = start_row_index

            # 设置样式
            black_color = '000000'
            white_color = 'FFFFFF'
            header_font = Font('微软雅黑', size=12, bold=True, color=white_color)
            font = Font('微软雅黑', size=12, color=black_color)
            header_fill = PatternFill('solid', fgColor='FF5722')
            header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            alignment = Alignment(vertical='center', wrap_text=True)
            header_border = Border(left=Side(border_style='thin', color=white_color),
                                   right=Side(border_style='thin', color=white_color),
                                   top=Side(border_style='thin', color=white_color),
                                   bottom=Side(border_style='thin', color=white_color))
            border = Border(left=Side(border_style='thin', color=black_color),
                            right=Side(border_style='thin', color=black_color),
                            top=Side(border_style='thin', color=black_color),
                            bottom=Side(border_style='thin', color=black_color))

            worksheet['A1'].font = header_font
            worksheet['B1'].font = header_font
            worksheet['C1'].font = header_font
            worksheet['D1'].font = header_font
            worksheet['E1'].font = header_font

            worksheet['A1'].fill = header_fill
            worksheet['B1'].fill = header_fill
            worksheet['C1'].fill = header_fill
            worksheet['D1'].fill = header_fill
            worksheet['E1'].fill = header_fill

            worksheet['A1'].border = header_border
            worksheet['B1'].border = header_border
            worksheet['C1'].border = header_border
            worksheet['D1'].border = header_border
            worksheet['E1'].border = header_border

            worksheet['A1'].alignment = header_alignment
            worksheet['B1'].alignment = header_alignment
            worksheet['C1'].alignment = header_alignment
            worksheet['D1'].alignment = header_alignment
            worksheet['E1'].alignment = header_alignment

            # 第一行
            worksheet.cell(row=1, column=default_index_col).value = '序号'
            worksheet.cell(row=1, column=default_word_col).value = '单词'
            worksheet.cell(row=1, column=default_phonetic_col).value = '音标'
            worksheet.cell(row=1, column=default_translation_col).value = '翻译'
            worksheet.cell(row=1, column=default_example_col).value = '例句'

            for word, phonetic, translate, example in parse(file_path):
                worksheet['A{}'.format(current_row_index)].font = font
                worksheet['B{}'.format(current_row_index)].font = font
                worksheet['C{}'.format(current_row_index)].font = font
                worksheet['D{}'.format(current_row_index)].font = font
                worksheet['E{}'.format(current_row_index)].font = font

                worksheet['A{}'.format(current_row_index)].border = border
                worksheet['B{}'.format(current_row_index)].border = border
                worksheet['C{}'.format(current_row_index)].border = border
                worksheet['D{}'.format(current_row_index)].border = border
                worksheet['E{}'.format(current_row_index)].border = border

                worksheet['A{}'.format(current_row_index)].alignment = header_alignment
                worksheet['B{}'.format(current_row_index)].alignment = alignment
                worksheet['C{}'.format(current_row_index)].alignment = alignment
                worksheet['D{}'.format(current_row_index)].alignment = alignment
                worksheet['E{}'.format(current_row_index)].alignment = alignment

                worksheet.cell(row=current_row_index, column=default_index_col).value = current_row_index - 1
                worksheet.cell(row=current_row_index, column=default_word_col).value = word
                worksheet.cell(row=current_row_index, column=default_phonetic_col).value = phonetic
                worksheet.cell(row=current_row_index, column=default_translation_col).value = translate
                worksheet.cell(row=current_row_index, column=default_example_col).value = example
                # 下一行
                current_row_index += 1
            workbook.save(xlsx_path)
            workbook.close()


def parse(file_path):
    with open(file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    for line in lines:
        json_data = json.loads(line)
        word = json_data['content']['word']['wordHead']
        content = json_data['content']['word']['content']
        # 例句
        example = ''
        sentence = content.get('sentence')
        if sentence:
            sentences = sentence.get('sentences')
            if sentences:
                for s in sentences:
                    example += '{}({})\n'.format(s.get('sContent'), s.get('sCn'))
        example = example.strip()
        # 音标
        phonetic = ''
        ukphone = content.get('ukphone')
        usphone = content.get('usphone')
        if ukphone:
            phonetic += '英/{}/ '.format(ukphone)
        if usphone:
            phonetic += '美/{}/ '.format(usphone)
        phonetic = phonetic.strip()
        # 翻译
        translate = ''
        syno = content.get('syno')
        if syno:
            synos = syno.get('synos')
            if synos:
                for s in synos:
                    translate += '{}.{} '.format(s.get('pos'), s.get('tran'))
        translate = translate.strip()
        yield word, phonetic, translate, example


if __name__ == '__main__':
    worksheet_name = 'Sheet1'
    # 开始行标，默认从第2行开始读
    start_row_index = 2
    # 默认序号列
    default_index_col = 1
    # 默认单词列
    default_word_col = 2
    # 默认音标列
    default_phonetic_col = 3
    # 默认翻译列
    default_translation_col = 4
    # 默认例句列
    default_example_col = 5
    # parse_all('.\\book', '.\\excel')
    re_order_index('.\\excel')
